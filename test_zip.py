import zipfile
import pytest
from pypdf import PdfReader
import csv
from openpyxl import load_workbook

from conftest import TMP_DIR


def test_check_csv(new_archive):
    with zipfile.ZipFile(TMP_DIR + "/tmp_zip_test.zip") as zip_file:
        with zip_file.open("CSV.csv") as csv_file:
            content = csv_file.read().decode('utf-8')
            csvreader = list(csv.reader(content.splitlines(), delimiter=','))
            second_row = csvreader[1]
            assert second_row[0] == 'Edyth Berge'
            assert second_row[1] == '414-680-5073'


def test_check_pdf(new_archive):
    with zipfile.ZipFile(TMP_DIR + "/tmp_zip_test.zip") as zip_file:
        with zip_file.open("PDF.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            assert len(reader.pages) == 256
            assert "Python Testing with pytest" in reader.pages[1].extract_text()


def test_check_xlsx(new_archive):
    with zipfile.ZipFile(TMP_DIR + "/tmp_zip_test.zip") as zip_file:
        with zip_file.open("XLSX.xlsx") as xlsx_file:
            content = load_workbook(xlsx_file)
            sheet = content.active
            assert sheet.cell(row=2, column=2).value == "Dulce"
            assert sum(1 for _ in sheet.iter_rows()) == 51
            assert sum(1 for _ in sheet.iter_cols()) == 8
